from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
import sys
import tempfile
import cv2
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QDialog, QLineEdit, QSizePolicy, QApplication, QMainWindow, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QWidget, QComboBox, QMessageBox
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtGui import QImage, QPixmap
from fer import FER
from deepface import DeepFace
import sqlite3
from datetime import datetime, timedelta
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from tkcalendar import DateEntry  # Usando tkcalendar para el selector de fechas
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import tensorflow as tf
import io
from datetime import date
from email.mime.base import MIMEBase
from email import encoders
import zipfile, os, io
from datetime import date
import numpy as np

# Base de datos SQLite
DB_NAME = "emociones.db"

DEBUG_DIR = "debug_faces"
os.makedirs(DEBUG_DIR, exist_ok=True)

# Umbral empírico para reconocer
FACE_THRESHOLD = 0.90

# { "José": np.array([...]), "María": np.array([...]), … }
known_embeddings = {}

# Inicializar la base de datos
def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # 1) Persona: solo guarda el nombre
    c.execute('''
      CREATE TABLE IF NOT EXISTS personas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE
      )
    ''')
    # 2) Rostros: una fila por cada imagen capturada
    c.execute('''
      CREATE TABLE IF NOT EXISTS rostros_persona (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        persona_id INTEGER,
        image_path TEXT,
        FOREIGN KEY(persona_id) REFERENCES personas(id)
      )
    ''')

    c.execute('''CREATE TABLE IF NOT EXISTS personas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE,
        image_path TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS emociones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        emocion TEXT,
        porcentaje REAL,
        fecha_hora TEXT,
        fecha TEXT,
        imagen_path TEXT
    )''')
    
    # Nueva tabla de usuarios
    c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT UNIQUE,
        contrasena TEXT
    )''')
    
    # Crear usuario por defecto si no existe
    c.execute("SELECT * FROM usuarios WHERE usuario = 'admin'")
    if not c.fetchone():
        c.execute("INSERT INTO usuarios (usuario, contrasena) VALUES (?, ?)", ("admin", "admin123"))
        
    # Añadir columna nombre si no existe
    c.execute("PRAGMA table_info(emociones)")
    cols = [r[1] for r in c.fetchall()]
    if "nombre" not in cols:
        c.execute("ALTER TABLE emociones ADD COLUMN nombre TEXT")

    conn.commit()
    conn.close()
    
def mostrar_aplicacion():
    # Aquí va la lógica principal de la aplicación una vez logueado
    ventana = tk.Tk()
    ventana.title("Aplicación Principal")
    ventana.geometry("400x300")
    tk.Label(ventana, text="¡Bienvenido!", font=("Helvetica", 16)).pack(pady=50)
    ventana.mainloop()
    
def mostrar_login():
    def verificar_credenciales(event=None):
        usuario = entry_usuario.get()
        contrasena = entry_contrasena.get()

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT * FROM usuarios WHERE usuario = ? AND contrasena = ?", (usuario, contrasena))
        resultado = c.fetchone()
        conn.close()

        if resultado:
            login_window.destroy()
            mostrar_aplicacion()
        else:
            messagebox.showerror("Error", "Credenciales incorrectas")

    login_window = tk.Tk()
    login_window.title("Inicio de Sesión")
    login_window.resizable(False, False)
    login_window.configure(bg="#f0f4f7")

    # Tamaño de la ventana
    ancho_ventana = 350
    alto_ventana = 450

    # Obtener dimensiones de la pantalla
    ancho_pantalla = login_window.winfo_screenwidth()
    alto_pantalla = login_window.winfo_screenheight()

    # Calcular posición x, y
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)

    # Posicionar la ventana en el centro
    login_window.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    # Contenedor visual
    frame = tk.Frame(login_window, bg="white", padx=80, pady=80, relief="ridge", bd=2)
    frame.place(relx=0.5, rely=0.5, anchor="center")

    tk.Label(frame, text="Inicia Sesión", font=("Helvetica", 16), bg="white").pack(pady=(0, 10))

    tk.Label(frame, text="Usuario:", bg="white", anchor="w").pack(fill='x')
    entry_usuario = tk.Entry(frame)
    entry_usuario.pack(fill='x', pady=(0, 10))
    entry_usuario.focus()  # Foco automático

    tk.Label(frame, text="Contraseña:", bg="white", anchor="w").pack(fill='x')
    entry_contrasena = tk.Entry(frame, show="*")
    entry_contrasena.pack(fill='x', pady=(0, 10))

    tk.Button(frame, text="Ingresar", bg="#4CAF50", fg="white", command=verificar_credenciales).pack(pady=(10, 0), fill='x')

    login_window.bind("<Return>", verificar_credenciales)

    login_window.mainloop()

# Guardar resultados en SQLite
def guardar_resultado_sqlite(emocion, porcentaje):
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT INTO emociones (emocion, porcentaje, fecha_hora, fecha) VALUES (?, ?, ?, ?)",
              (emocion, round(porcentaje * 100, 2), fecha_hora, fecha))
    conn.commit()
    conn.close()
    
# def guardar_resultado_sqlitetest():
#     fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#     fecha = datetime.now().strftime("2025-05-04")
#     conn = sqlite3.connect(DB_NAME)
#     c = conn.cursor()
#     emociones = ["triste", "miedo", "enojo"]

#     for i in range(30):
#         emocion = random.choice(emociones)
#         confianza = random.uniform(50, 100)  # Rango de confianza entre 50% y 100%

#         if confianza > 70:
#             print(f"FER (alta confianza): {emocion} ({confianza:.2f}%)")
#             c.execute("INSERT INTO emociones (emocion, porcentaje, fecha_hora, fecha) VALUES (?, ?, ?, ?)",
#                     (emocion, round(confianza, 2), fecha_hora, fecha))
#     conn.commit()
#     conn.close()


def guardar_resultado_con_imagen(emocion, porcentaje, imagen_path=None,nombre=None):
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        INSERT INTO emociones 
          (emocion, porcentaje, fecha_hora, fecha, imagen_path, nombre)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (emocion, round(porcentaje * 100, 2), fecha_hora, fecha, imagen_path, nombre))
    conn.commit()
    conn.close()

# Función para buscar emociones por fecha
def buscar_emociones_por_fecha(inicio, fin):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    # Convertir las fechas a formato de cadena compatible con SQLite
    inicio_str = inicio.strftime('%Y-%m-%d')
    fin_str = fin.strftime('%Y-%m-%d')
    
    c.execute("SELECT emocion, porcentaje, fecha_hora FROM emociones WHERE imagen_path is NULL and fecha BETWEEN ? AND ?", (inicio_str, fin_str))
    resultados = c.fetchall()
    conn.close()
    return resultados

# Función para obtener la lista de emociones desde la base de datos
def obtener_lista_emociones():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT emocion, porcentaje, fecha_hora FROM emociones")
    resultados = c.fetchall()
    conn.close()
    return resultados

# Función para generar un archivo Excel con la lista de emociones
def generar_excel_emociones():
    try:
        # Obtener la lista de emociones
        lista_emociones = obtener_lista_emociones()
        if not lista_emociones:
            messagebox.showwarning("Advertencia", "No hay emociones registradas en la base de datos.")
            return None

        # Crear un DataFrame de pandas
        df = pd.DataFrame(lista_emociones, columns=["Emoción", "Porcentaje", "FechaHora"])

        # Guardar el archivo Excel temporalmente
        archivo_excel = "emociones.xlsx"
        df.to_excel(archivo_excel, index=False)
        return archivo_excel
    except Exception as e:
        messagebox.showerror("Error", f"Error al generar el archivo Excel: {str(e)}")
        return None

def enviar_excel_emociones_por_correo():
    try:
        # Generar el archivo Excel
        archivo_excel = generar_excel_emociones()
        if not archivo_excel:
            return

        # Configurar el correo
        subject = "Lista de Emociones Registradas"
        body = "Adjunto encontrarás un archivo Excel con la lista de emociones registradas."
        to_email = "gianella.taboada@gmail.com"  # Cambia esto al correo deseado

        from_email = "richardalvarezruiz.1997@gmail.com"  # Tu correo de Gmail
        password = "kobh lpxw mzcf vwqb"  # Tu contraseña de Gmail o contraseña de aplicación

        # Crear el mensaje de correo
        mensaje = MIMEMultipart()
        mensaje["From"] = from_email
        mensaje["To"] = to_email
        mensaje["Subject"] = subject
        mensaje.attach(MIMEText(body, "plain"))

        # Adjuntar el archivo Excel
        with open(archivo_excel, "rb") as adjunto:
            part = MIMEText(adjunto.read(), "base64", "utf-8")
            part.add_header("Content-Disposition", f"attachment; filename={archivo_excel}")
            mensaje.attach(part)

        # Enviar el correo
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(from_email, password)
        server.sendmail(from_email, to_email, mensaje.as_string())
        server.quit()

        messagebox.showinfo("Éxito", "El archivo Excel se ha enviado por correo.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al enviar el archivo Excel: {str(e)}")

# Configuración para guardar las imágenes
image_save_folder = "capturas_tristeza"  # Carpeta donde se guardarán las capturas
if not os.path.exists(image_save_folder):
    os.makedirs(image_save_folder)  # Crear la carpeta si no existe

# Función para guardar la imagen cuando se detecta la emoción
def guardar_imagen_tristeza(frame):
    # Crear un nombre único para la imagen con la fecha y hora actual
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    image_filename = os.path.join(image_save_folder, f"tristeza_{timestamp}.jpg")
    
    # Guardar la imagen
    cv2.imwrite(image_filename, frame)
    print(f"Imagen guardada como {image_filename}")
    return image_filename

# 1) Definir una función genérica:
def guardar_imagen_emocion(face_img, emocion, label_persona):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Si no se reconoció, marca como "Desconocido"
    persona = label_persona if label_persona else "Desconocido"

    # Carpeta donde guardas
    folder = "emociones_negativas"
    os.makedirs(folder, exist_ok=True)

    # Construye el nombre incluyendo persona y emoción
    filename = f"{persona}_{emocion}_{timestamp}.jpg"
    path = os.path.join(folder, filename)

    cv2.imwrite(path, face_img)
    return path

def recognize_face(face_img_bgr):
    # 0) Guarda la imagen de depuración
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    debug_path = os.path.join(DEBUG_DIR, f"compare_{ts}.jpg")
    cv2.imwrite(debug_path, face_img_bgr)
    print(f"[Reconocimiento] guardada imagen de comparación en {debug_path}")

    # 1) Guardar en temp para DeepFace…
    fd, tmp_path = tempfile.mkstemp(suffix=".jpg")
    os.close(fd)
    cv2.imwrite(tmp_path, face_img_bgr)

    # 2) Obtener embedding
    rep_list = DeepFace.represent(
        img_path=tmp_path,
        model_name='Facenet',
        enforce_detection=False,
        detector_backend='mtcnn'
    )
    os.remove(tmp_path)
    vec = np.array(rep_list[0]["embedding"], dtype=float)

    # 1) Normalizar el vector de prueba
    norm_vec = np.linalg.norm(vec)
    if norm_vec > 0:
        vec = vec / norm_vec
    print(f"[recognize_face] norma vector prueba: {norm_vec:.4f}")

    # 2) Comparar contra cada known_vec (que ya está normalizado)
    best_match, min_dist = None, float("inf")
    for nombre, known_vec in known_embeddings.items():
        dist = np.linalg.norm(known_vec - vec)
        print(f"   → distancia a «{nombre}»: {dist:.4f}")
        if dist < min_dist:
            min_dist, best_match = dist, nombre

    print(f"   >>> mejor match: {best_match!r} con dist={min_dist:.4f} (umbral={FACE_THRESHOLD})")
    return best_match if min_dist <= FACE_THRESHOLD else None

# Función para exportar los resultados a un archivo CSV
def exportar_a_csv():
    try:
        # Obtener el texto de la caja de resultados
        data = resultado_text.get("1.0", tk.END).strip().split("\n")
        
        if not data:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")
            return
        
        # Abrir un cuadro de diálogo para que el usuario elija la ubicación y el nombre del archivo
        archivo_guardado = filedialog.asksaveasfilename(
            defaultextension=".csv", 
            filetypes=[("Archivos CSV", "*.csv")],
            title="Guardar archivo CSV"
        )
        
        if archivo_guardado:  # Si el usuario elige una ubicación
            # Abrir el archivo CSV en modo de escritura
            with open(archivo_guardado, mode="w", newline='') as archivo:
                writer = csv.writer(archivo)
                
                # Escribir los encabezados del CSV
                writer.writerow(["Emoción", "Porcentaje", "FechaHora"])
                
                # Escribir cada línea de datos (ignoramos la primera línea vacía o de encabezado)
                for line in data:
                    if line:
                        parts = line.split(", ")
                        emocion = parts[0].replace("Emoción: ", "")
                        porcentaje = parts[1].replace("Porcentaje: ", "").replace("%", "")
                        fecha_hora = parts[2].replace("Fecha: ", "")
                        writer.writerow([emocion, porcentaje, fecha_hora])
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente.")
        else:
            messagebox.showwarning("Advertencia", "No se seleccionó ninguna ubicación para guardar el archivo.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar los datos: {str(e)}")
        
def exportar_a_excel():
    try:
        # Obtener el texto de la caja de resultados
        data = resultado_text.get("1.0", tk.END).strip().split("\n")
        
        if not data:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")
            return
        
        # Preparar los datos para el DataFrame de pandas
        rows = []
        for line in data:
            if line:
                parts = line.split(", ")
                emocion = parts[0].replace("Emoción: ", "")
                porcentaje = parts[1].replace("Porcentaje: ", "").replace("%", "")
                fecha_hora = parts[2].replace("Fecha: ", "")
                rows.append([emocion, porcentaje, fecha_hora])
        
        # Crear un DataFrame de pandas
        df = pd.DataFrame(rows, columns=["Emoción", "Porcentaje", "FechaHora"])

        # Abrir un cuadro de diálogo para que el usuario elija la ubicación y el nombre del archivo
        archivo_guardado = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar archivo Excel"
        )
        
        if archivo_guardado:  # Si el usuario elige una ubicación
            # Crear un libro de trabajo (workbook) con openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Emociones"

            # Escribir los encabezados de columna en mayúsculas y con color de fondo
            headers = ["Emoción", "Porcentaje", "FechaHora"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header.upper())  # Títulos en mayúsculas
                cell.font = Font(bold=True)  # Negrita
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo

            # Escribir los datos en las celdas
            for row_num, row in enumerate(df.itertuples(index=False, name=None), 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

            # Guardar el archivo Excel
            wb.save(archivo_guardado)
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente.")
        else:
            messagebox.showwarning("Advertencia", "No se seleccionó ninguna ubicación para guardar el archivo.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar los datos: {str(e)}")

# Iniciar la interfaz de búsqueda de fechas con Tkinter
def iniciar_gui_tkinter():
    def buscar_por_fecha():
        try:
            # Convertir las entradas a tipo datetime
            fecha_inicio = pd.to_datetime(entry_inicio.get())
            fecha_fin = pd.to_datetime(entry_fin.get())

            # Llamar la función para obtener los resultados
            resultados = buscar_emociones_por_fecha(fecha_inicio, fecha_fin)

            resultado_text.delete("1.0", tk.END)  # Limpiar el campo de texto antes de mostrar los nuevos resultados
            if not resultados:
                resultado_text.insert(tk.END, "No hay resultados en ese rango.")
            else:
                # Mostrar los resultados
                for row in resultados:
                    emocion, porcentaje, fecha_hora = row
                    resultado_text.insert(tk.END, f"Emoción: {emocion}, Porcentaje: {porcentaje}%, Fecha: {fecha_hora}\n")
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar las fechas.\n{str(e)}")

    ventana = tk.Tk()
    ventana.title("Buscar emociones por fecha")
    ventana.geometry("600x500")

    ttk.Label(ventana, text="Fecha Inicio:").pack(pady=5)
    global entry_inicio
    entry_inicio = DateEntry(ventana, date_pattern='yyyy-mm-dd')
    entry_inicio.pack()

    ttk.Label(ventana, text="Fecha Fin:").pack(pady=5)
    global entry_fin
    entry_fin = DateEntry(ventana, date_pattern='yyyy-mm-dd')
    entry_fin.pack()

    ttk.Button(ventana, text="Buscar", command=buscar_por_fecha).pack(pady=10)

    # Crear un frame para contener la barra de desplazamiento y el widget de texto
    frame = ttk.Frame(ventana)
    frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

    # Crear el widget de Scrollbar
    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Crear el widget de Text
    global resultado_text
    resultado_text = tk.Text(frame, height=15, width=70)
    resultado_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Asociar el Scrollbar con el Text widget
    scrollbar.config(command=resultado_text.yview)
    resultado_text.config(yscrollcommand=scrollbar.set)

    # Crear el botón de Exportar a CSV
    # export_button = ttk.Button(ventana, text="Exportar a CSV", command=exportar_a_csv)
    export_button = ttk.Button(ventana, text="Exportar a EXCEL", command=exportar_a_excel)
    export_button.pack(pady=10)

    ventana.mainloop()
    
# Función para enviar correo electrónico
def enviar_correo(subject, body, to_email):
    # Configuración de correo
    from_email = "richardalvarezruiz.1997@gmail.com"  # Tu correo de Gmail
    password = "kobh lpxw mzcf vwqb"  # Tu contraseña de Gmail o contraseña de aplicación

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(from_email, password)

        mensaje = MIMEMultipart()
        mensaje["From"] = from_email
        mensaje["To"] = to_email
        mensaje["Subject"] = subject
        mensaje.attach(MIMEText(body, "plain"))

        server.sendmail(from_email, to_email, mensaje.as_string())
        print("Correo enviado correctamente.")
    except smtplib.SMTPAuthenticationError as e:
            print("Error de autenticación SMTP:", e)
    except Exception as e:
            print("Error enviando el correo:", e)
    finally:
            server.quit()
    
def verificar_y_enviar_correo(emotion, porcentaje):
    if emotion in ['sad', 'fear', 'angry', 'disgust']:
        subject = "⚠️ Alerta: Emoción Negativa Detectada"
        body = f"Se ha detectado la emoción de: {emotion.capitalize()}.\n\nPorcentaje: {porcentaje * 100:.2f}%"
        to_email = "gianella.taboada@gmail.com"  # Correo a donde se enviará la alerta
        enviar_correo(subject, body, to_email)
        print(f"Correo enviado: {subject}")

# Clase para mostrar gráfico de emociones
class EmotionHistoryPlot(QWidget):
    def __init__(self): 
        super().__init__()
        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)
        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.setLayout(layout)

    def update_plot(self, emotions_count):
        self.ax.clear()
        emotions = list(emotions_count.keys())
        counts = list(emotions_count.values())

        self.ax.bar(emotions, counts, color='skyblue')
        self.ax.set_ylabel('Frecuencia')
        self.ax.set_title('Historial de Emociones')

        for i, count in enumerate(counts):
            self.ax.text(i, count, f'({count})', ha='center', va='bottom')

        self.canvas.draw()

# Clase principal de la aplicación PyQt5
class DepressionDetector(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Detección de Conductas Depresivas")
        self.setGeometry(200, 200, 1500, 600)

        self.video_label = QLabel(self)
        self.video_label.setAlignment(Qt.AlignCenter)
        
        # justo después de crear self.video_label:
        self.video_label.setSizePolicy(
            QSizePolicy.Expanding,
            QSizePolicy.Expanding
        )
        # opcional: ponle un mínimo razonable para que no colapse
        self.video_label.setMinimumSize(320*2, 240*2)

        self.start_button = QPushButton("Iniciar Detección", self)
        self.start_button.clicked.connect(self.start_detection)
       
        self.search_button = QPushButton("Buscar por Fecha", self)
        self.search_button.clicked.connect(self.abrir_busqueda_por_fecha)
        
        # Botón para enviar el archivo Excel por correo
        self.send_excel_button = QPushButton("Enviar Lista de Emociones en Excel", self)
        self.send_excel_button.clicked.connect(self.enviar_excel_emociones)

        self.model_selector = QComboBox(self)
        self.model_selector.addItem("FER")
        self.model_selector.addItem("DeepFace")
        
        #Botón de información
        self.info_button = QPushButton("?")
        self.info_button.setFixedSize(25, 25)
        self.info_button.setToolTip("Información sobre los modelos")
        self.info_button.clicked.connect(self.mostrar_info_modelos)
        
        # # Estilo circular del botón
        self.info_button.setStyleSheet("""
            QPushButton {
                border-radius: 12px;
                font-weight: bold;
                background-color: #3498db;
                color: white;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        model_row_layout = QHBoxLayout()
        model_row_layout.addWidget(self.model_selector, 9)
        model_row_layout.addWidget(self.info_button, 1) 
        
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("Selecciona el modelo:"))
        left_layout.addLayout(model_row_layout)
        left_layout.addWidget(self.video_label)
        
        left_layout.addStretch()
        
        left_layout.addWidget(self.start_button)
        left_layout.addWidget(self.search_button)
        left_layout.addWidget(self.send_excel_button)
        
        stats_layout = QVBoxLayout()
        self.dominant_emotion_label = QLabel("Emociones Dominantes: ", self)
        stats_layout.addWidget(self.dominant_emotion_label)

        self.average_emotion_label = QLabel("Promedio de Emociones: ", self)
        stats_layout.addWidget(self.average_emotion_label)

        self.alert_label = QLabel("", self)
        stats_layout.addWidget(self.alert_label)

        self.emotion_history_plot = EmotionHistoryPlot()
        stats_layout.addWidget(self.emotion_history_plot)

        main_layout = QHBoxLayout()
        main_layout.addLayout(left_layout)
        main_layout.addLayout(stats_layout)
        

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)
        
        self.cap = cv2.VideoCapture(0)
        #self.cap = cv2.VideoCapture("video_prueba_2.mp4") # Cambia el índice a 0 para hacerlo con una camara
        # self.cap = cv2.VideoCapture("video2.mp4") # Cambia el índice por el nombre del video "video2.mp4" para reconocer un video
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)

        self.all_list_emotions = {
            'sad': 'Tristeza', 'fear': 'Miedo', 'angry': 'Enojo', 'disgust': 'Desprecio',
            'happy': 'Feliz', 'surprise': 'Sorpresa', 'neutral': 'Neutral'
        }
        self.depressive_emotions = {
            'sad': 'Tristeza', 'fear': 'Miedo', 'angry': 'Enojo', 'disgust': 'Desprecio'
        }
        
        self.ultima_captura_tristeza = datetime.min  # Inicializa con la fecha mínima posible
        self.intervalo_captura = timedelta(seconds=60) # Intervalo de tiempo en segundos entre capturas

        self.fer_detector = FER(mtcnn=True)
        self.emotion_history = []
        self.emotions_count = {emotion: 0 for emotion in self.all_list_emotions.values()}
        self.negative_emotion_history = []
        self.max_history_size = 100
        self.emocion_anterior = None

        # 1) Intervalo configurable (en minutos)
        self.report_interval_minutes = 30

        # 2) Lista de destinatarios
        self.recipients = [
            "gianella.taboada@gmail.com"
        ]        
        
        # 5) Timer automático cada 30 minutos
        self.half_hour_timer = QTimer(self)
        self.half_hour_timer.timeout.connect(self.send_half_hour_report)
        # 30 min = 30*60*1000 ms
        self.half_hour_timer.start(self.report_interval_minutes * 60 * 1000)        

        # 2) Creamos un ThreadPool para no bloquear la UI
        self.executor = ThreadPoolExecutor(max_workers=1)
        self.processing = False      # bandera
        self.last_frame = None       # último frame procesado
        threading.Thread(target=self._warmup_deepface, daemon=True).start()

        # main.py (añadir dentro de __init__ de DepressionDetector) :contentReference[oaicite:0]{index=0}
        self.daily_report_button = QPushButton("Enviar reporte diario", self)
        self.daily_report_button.clicked.connect(self.send_daily_report)
        left_layout.insertWidget(left_layout.indexOf(self.send_excel_button)+1, self.daily_report_button)

        self.register_face_button = QPushButton("Registrar Rostro", self)
        self.register_face_button.clicked.connect(self.open_face_registration)
        left_layout.insertWidget(
            left_layout.indexOf(self.daily_report_button), 
            self.register_face_button
        )


    def open_face_registration(self):
        # 1) parar y liberar en main
        self.timer.stop()
        self.cap.release()

        # 2) abrir diálogo (que abrirá su propio VideoCapture)
        dialog = FaceRegistrationDialog(self)
        dialog.exec_()

        # 3) al cerrar diálogo, vuelves a inicializar main
        self.cap = cv2.VideoCapture(0)
        self.timer.start(30)    # el mismo intervalo que usabas

    def generate_daily_excel_report(self):
        conn = sqlite3.connect(DB_NAME)
        today = date.today().strftime("%Y-%m-%d")
        c = conn.cursor()
        # Conteo por emoción
        c.execute("SELECT emocion, COUNT(*) FROM emociones WHERE fecha=? GROUP BY emocion", (today,))
        data = c.fetchall()
        # Otras métricas
        c.execute("SELECT COUNT(*) FROM emociones WHERE fecha=?", (today,))
        total = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM emociones WHERE fecha=? AND imagen_path IS NOT NULL", (today,))
        images = c.fetchone()[0]
        conn.close()

        # Crear DataFrame
        df = pd.DataFrame(data, columns=["Emoción", "Cantidad"])
        # Añadir totales al Excel
        df_totals = pd.DataFrame({
            "Métrica": ["Total registros", "Capturas imágenes"],
            "Valor": [total, images]
        })
        # Escribir a Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Por emoción", index=False)
            df_totals.to_excel(writer, sheet_name="Resumen", index=False)
        output.seek(0)
        return output  # BytesIO con el archivo Excel

    def send_daily_report(self):
        self.send_half_hour_report()

    def compose_report_html(self, start: datetime, end: datetime, total: int, images: int, data: list):
        # Construye filas de distribución
        rows_html = "\n".join(
            f"<tr><td>{emocion}</td><td>{cantidad}</td></tr>"
            for emocion, cantidad in data
        )
        html = f"""
        <html>
        <body>
            <h2>Reporte emociones {start.strftime('%Y-%m-%d %H:%M')} – {end.strftime('%Y-%m-%d %H:%M')}</h2>
            <p><strong>Total registros:</strong> {total} &nbsp; | &nbsp;
               <strong>Capturas imágenes:</strong> {images}</p>
            <p><strong>Distribución por emoción:</strong></p>
            <table border="1" cellpadding="4" cellspacing="0">
                <tr><th>Emoción</th><th>Cantidad</th></tr>
                {rows_html}
            </table>
            <p>Adjunto encontrarás el detalle en Excel y las imágenes capturadas.</p>
        </body>
        </html>
        """
        return html
    
    def send_half_hour_report(self):
        def _worker():
            end = datetime.now()
            start = end - timedelta(minutes=self.report_interval_minutes)
            excel_buf, total, images, data = self.generate_interval_report(start, end)
            zip_buf = self.generate_images_zip(start, end)

            subject = f"Reporte emociones {self.report_interval_minutes}′ – {end.strftime('%Y-%m-%d %H:%M')}"
            to_email = ", ".join(self.recipients)
            body_html = self.compose_report_html(start, end, total, images, data)

            msg = MIMEMultipart()
            msg["From"] = from_email = "richardalvarezruiz.1997@gmail.com"
            msg["To"] = to_email
            msg["Subject"] = subject
            msg.attach(MIMEText(body_html, "html"))

            # Adjuntar Excel
            part_xl = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part_xl.set_payload(excel_buf.read())
            encoders.encode_base64(part_xl)
            part_xl.add_header("Content-Disposition", f'attachment; filename="intervalo_{end.strftime("%Y%m%d_%H%M")}.xlsx"')
            msg.attach(part_xl)

            # Adjuntar ZIP
            part_zip = MIMEBase("application", "zip")
            part_zip.set_payload(zip_buf.read())
            encoders.encode_base64(part_zip)
            part_zip.add_header("Content-Disposition", f'attachment; filename="imgs_{end.strftime("%Y%m%d_%H%M")}.zip"')
            msg.attach(part_zip)

            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(from_email, "kobh lpxw mzcf vwqb")
            server.send_message(msg)
            server.quit()

            # Feedback en la UI
            self.alert_label.setText(f"Reporte 30′ enviado a {to_email}")
        threading.Thread(target=_worker, daemon=True).start()    
    
    def generate_images_zip(self, start: datetime, end: datetime):
            """
            Genera un ZIP con las imágenes cuyo `fecha_hora` esté entre start y end.
            """
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            # Filtramos por datetime(fecha_hora)
            c.execute(
                """
                SELECT imagen_path
                FROM emociones
                WHERE datetime(fecha_hora) BETWEEN ? AND ?
                AND imagen_path IS NOT NULL
                """,
                (
                    start.strftime("%Y-%m-%d %H:%M:%S"),
                    end.strftime("%Y-%m-%d %H:%M:%S")
                )
            )
            paths = [row[0] for row in c.fetchall()]
            conn.close()

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for img_path in paths:
                    if os.path.exists(img_path):
                        zipf.write(img_path, arcname=os.path.basename(img_path))
            zip_buffer.seek(0)
            return zip_buffer
    
    def generate_interval_report(self, start: datetime, end: datetime):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        start_str = start.strftime("%Y-%m-%d %H:%M:%S")
        end_str   = end.strftime("%Y-%m-%d %H:%M:%S")

        # 1) Conteo por emoción
        c.execute("""
            SELECT emocion, COUNT(*) 
            FROM emociones 
            WHERE datetime(fecha_hora) BETWEEN ? AND ?
            GROUP BY emocion
        """, (start_str, end_str))
        data = c.fetchall()

        # 2) Totales
        c.execute("""
            SELECT COUNT(*) 
            FROM emociones 
            WHERE datetime(fecha_hora) BETWEEN ? AND ?
        """, (start_str, end_str))
        total = c.fetchone()[0]

        c.execute("""
            SELECT COUNT(*) 
            FROM emociones 
            WHERE datetime(fecha_hora) BETWEEN ? AND ?
            AND imagen_path IS NOT NULL
        """, (start_str, end_str))
        images = c.fetchone()[0]

        # 3) Reconocimientos: traemos todas las filas
        c.execute("""
            SELECT fecha_hora AS Hora, nombre AS Persona, emocion AS Emoción
            FROM emociones
            WHERE datetime(fecha_hora) BETWEEN ? AND ?
            AND nombre IS NOT NULL
        """, (start_str, end_str))
        rows_rec = c.fetchall()                # <-- todas las filas
        conn.close()

        # 4) DataFrames
        df = pd.DataFrame(data, columns=["Emoción", "Cantidad"])
        df_totals = pd.DataFrame({
            "Métrica": ["Total registros", "Capturas imágenes"],
            "Valor": [total, images]
        })
        # si hubo reconocimientos, rellena; si no, quedará vacío
        df_rec = pd.DataFrame(rows_rec, columns=["Hora", "Persona", "Emoción"])

        # 5) Escribir en Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Por emoción",    index=False)
            df_totals.to_excel(writer, sheet_name="Resumen", index=False)
            df_rec.to_excel(writer, sheet_name="Reconocimientos", index=False)
        output.seek(0)

        return output, total, images, data

    def _warmup_deepface(self):
        # toma un frame del cap, lo convierte y llama analyze para precargar el modelo
        ret, frame = self.cap.read()
        if not ret:
            return
        img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        try:
            DeepFace.analyze(
                img,
                actions=['emotion'],
                enforce_detection=False,
                detector_backend='mtcnn'
            )
        except:
            pass  # ignoramos errores, sólo es para cargar el modelo

    def start_detection(self):        
        self.timer.start(33) # Intervalo de actualización del frame

    def abrir_busqueda_por_fecha(self):
        threading.Thread(target=iniciar_gui_tkinter).start()
        
    def enviar_excel_emociones(self):
        threading.Thread(target=enviar_excel_emociones_por_correo).start()
        
    def mostrar_info_modelos(self):
        info_text = (
            "Modelos disponibles:\n\n"
            "• FER (Facial Expression Recognition):\n"
            "  FER se recomienda para situaciones que requieren un análisis rápido y eficiente de emociones básicas, especialmente cuando se necesita una respuesta inmediata.\n\n"
            "• DeepFace:\n"
            "  DeepFace es más adecuado para contextos en los que se requiere una evaluación más precisa y profunda de las emociones faciales, debido a su mayor capacidad de análisis."
        )

        QMessageBox.information(self, "Información sobre los Modelos", info_text)


    def update_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            return
        
        # Convertimos y opcionalmente reducimos resolución

        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        selected_model = self.model_selector.currentText()
        emotions_count = {emotion: 0 for emotion in self.all_list_emotions.values()}

        self.frame_counter = getattr(self, 'frame_counter', 0) + 1
        if self.model_selector.currentText() == "DeepFace":
            # Sólo lanzamos un nuevo análisis si el anterior ya terminó
            small = cv2.resize(rgb_image, (320, 240))
            if not self.processing:
                self.processing = True
                # Pasamos la versión pequeña para acelerars
                self.executor.submit(self._deepface_worker, rgb_image)

            # Mientras tanto mostramos el último resultado
            #display = self.last_frame if self.last_frame is not None else small
        else:
            faces = self.fer_detector.detect_emotions(rgb_image)
            if not faces:
                return

            for face in faces:
                (x, y, w, h) = face["box"]
                emociones = face["emotions"]
                emotion = max(emociones, key=emociones.get)
                porcentaje = emociones[emotion]

                if w < 50 or h < 50 or porcentaje < 0.60:
                    continue  # Ignorar regiones pequeñas o emociones con baja confianza

                color = (255, 0, 0) if emotion in self.depressive_emotions else (0, 255, 0)
                emocion_cast = self.depressive_emotions.get(emotion, emotion)

                cv2.rectangle(rgb_image, (x, y), (x + w, y + h), color, 2)
                cv2.putText(rgb_image, f"{emocion_cast}: {porcentaje*100:.2f}%", (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)                

                es_depresiva = emotion in ['sad', 'fear', 'angry', 'disgust']

                face_bgr = cv2.cvtColor(rgb_image, cv2.COLOR_RGB2BGR)[y:y+h, x:x+w]

                if es_depresiva:
                    ahora = datetime.now()
                    emocion_cambio = emotion != getattr(self, 'emocion_anterior', None)

                    # Solo si la emoción ha cambiado se actualiza la anterior
                    if emocion_cambio:
                        self.emocion_anterior = emotion

                        label = self.all_list_emotions.get(emotion, emotion)
                        #threading.Thread(target=guardar_resultado_sqlite, args=(label, porcentaje)).start()
                        #threading.Thread(target=verificar_y_enviar_correo, args=(emotion, porcentaje)).start()
                   
                        # Aumenta el contador solo si la emoción ha cambiado
                        if emotion in self.depressive_emotions:
                            label_persona = recognize_face(face_bgr)
                            print(f"[FER] Reconocido: {label_persona} para la emoción {emotion}")
                            # 1) Guardar la imagen y obtener la ruta
                            bgr_para_guardar = cv2.cvtColor(rgb_image, cv2.COLOR_RGB2BGR)
                            ruta = guardar_imagen_emocion(bgr_para_guardar, label, label_persona)                            
                            # 2) Lanzar hilo solo para la inserción en BD                            
                            threading.Thread(
                                target=guardar_resultado_con_imagen,
                                args=(label, porcentaje, ruta, label_persona),
                                daemon=True
                            ).start()

                            self.ultima_captura_tristeza = ahora            
                            emocion_cast = self.depressive_emotions[emotion]
                            emotions_count[emocion_cast] += 1
                            print(f"[FER] Emoción depresiva detectada: {emotion} con porcentaje {porcentaje*100:.2f}%")
                else:
                    # Si no es una emoción depresiva, se limpia la emoción anterior
                    self.emocion_anterior = None
                                        # Si es happy/surprise/neutral, contamos también
                    label_pos = self.all_list_emotions[emotion]      # p.ej. 'Feliz', 'Sorpresa' o 'Neutral'
                    self.emotions_count[label_pos] += 1

        # Mostrar en GUI
        # Dentro de update_frame, al final:
        if selected_model == "DeepFace":
            # elegimos el último frame con dibujo, o el pequeño crudo si aún no hay
            frame_to_show = self.last_frame if self.last_frame is not None else rgb_image
        else:
            frame_to_show = rgb_image

        display = cv2.resize(frame_to_show, (rgb_image.shape[1], rgb_image.shape[0]))

        h, w, _ = display.shape
        qimg = QImage(display.data, w, h, 3*w, QImage.Format_RGB888)
        pix  = QPixmap.fromImage(qimg)
        self.video_label.setPixmap(pix)

        for emotion, count in emotions_count.items():
            if emotion in self.emotions_count:
                self.emotions_count[emotion] += count

        if emotions_count:
            if emotions_count:
                # Ordenar emociones dominantes de mayor a menor
                emotions_to_remove = ['Neutral', 'Feliz','Sorpresa']  
                filtered_emotions = {k: v for k, v in self.emotions_count.items() if k not in emotions_to_remove}                
                sorted_emotions = dict(sorted(filtered_emotions.items(), key=lambda item: item[1], reverse=True))
                dominant_emotions_text = "\n".join([f"{emotion}: {count}" for emotion, count in sorted_emotions.items()])
                self.dominant_emotion_label.setText(f"Emociones Dominantes:\n{dominant_emotions_text}")

                # Promedio de emociones
                total_emotions_value = 0.0
                emotion_values = {
                    'Tristeza': -1.0,
                    'Miedo': -1.0,
                    'Enojo': -1.0,
                    'Desprecio': -1.0,
                    'Feliz':1.0,
                    'Sorpresa':1.0,
                    'Neutral': 0.0
                }
                # print("*******************")                  
                total = sum(self.emotions_count.values())
                total_value = sum(emotion_values[e] * cnt for e, cnt in self.emotions_count.items())
                avg = total_value / total if total>0 else 0

                self.average_emotion_label.setText(f"Promedio de Emociones: {avg:.2f}")

                self.negative_emotion_history.append(avg)
                # Si la lista supera el tamaño máximo, eliminamos el valor más antiguo
                if len(self.negative_emotion_history) > self.max_history_size:
                    self.negative_emotion_history.pop(0)                
                # Calcular el promedio de los últimos valores
                smoothed_negative_emotion = sum(self.negative_emotion_history) / len(self.negative_emotion_history)
                
                #print ("smoothed_negative_emotion: ",smoothed_negative_emotion)
                # Alerta
                if smoothed_negative_emotion < 0:
                    self.alert_label.setText("Alerta: Predominan emociones negativas.")
                else:
                    self.alert_label.setText("")

            # Actualizar el gráfico
            emotions_to_remove = ['Neutral', 'Feliz','Sorpresa']  
            filtered_emotions = {k: v for k, v in self.emotions_count.items() if k not in emotions_to_remove}                            
            self.emotion_history_plot.update_plot(filtered_emotions)

    def _deepface_worker(self, img):
    
        # Aseguramos que existe el atributo para llevar la última emoción
        if not hasattr(self, 'deepface_emocion_anterior'):
            self.deepface_emocion_anterior = None

        #Esto corre en un hilo distinto. `img` ya es pequeño.
    
        try:
            rgb_for_model = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

            # Usamos el modelo pre-cargado y GPU si la tienes disponible
            result = DeepFace.analyze(
                rgb_for_model,
                actions=['emotion'],
                enforce_detection=True,
                detector_backend='opencv', #opencv,mtcnn,retinaface
            )

            # Aseguramos tratar siempre una lista de caras
            faces = result if isinstance(result, list) else [result]
            out = img.copy()
            for face_info in faces:                
                region = face_info.get('region', {})
                x, y, w, h = region.get('x',0), region.get('y',0), region.get('w',0), region.get('h',0)
                face_bgr = img[y:y+h, x:x+w]
                if w < 50 or h < 50:
                    continue
                
                aspect = w / float(h)
                if aspect < 0.7 or aspect > 1.3:
                    continue
                if w*h < 5000:
                    continue

                # 1) Obtener emociones y confianza
                probs = face_info['emotion']              # e.g. {'happy': 34.5, 'neutral': 30.2, …}
                dom  = face_info['dominant_emotion']       # e.g. 'happy'
                conf = probs[dom] / 100.0                  # normalizar a [0,1]

                THRESHOLD = 0.70

                THRESHOLD_DEPRESSIVE = 0.75
                THRESHOLD_NEUTRAL   = 0.50

                if dom in self.depressive_emotions and conf >= THRESHOLD_DEPRESSIVE:
                    etiqueta = self.depressive_emotions[dom]
                elif probs['neutral']/100 >= THRESHOLD_NEUTRAL:
                    dom, conf = 'neutral', probs['neutral']/100
                    etiqueta = 'Neutral'
                else:
                    # si ni depresiva ni neutral supera mínimos, considerarla “sorpresa” o “happy”
                    # o simplemente saltarla según lo que prefieras
                    continue

                # 4) Dibujo y guardado como antes
                color = (255,0,0) if dom in self.depressive_emotions else (0,255,0)
                cv2.rectangle(img, (x,y), (x+w,y+h), color, 2)
                cv2.putText(
                    img,
                    f"{self.all_list_emotions.get(dom, dom)}: {conf*100:.2f}%",
                    (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.9, color, 2
                )
                
                #cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
                #cv2.putText(img, f"{self.all_list_emotions.get(dom, dom)}: {conf*100:.2f}%",
                #            cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)                

                # Etiqueta en español
                label = self.all_list_emotions.get(dom, dom)

                # Solo si la emoción cambió, la registramos
                if label != self.deepface_emocion_anterior:
                    self.deepface_emocion_anterior = label

                    # Guardar en BD en segundo plano
                    #threading.Thread(
                    #    target=guardar_resultado_sqlite,
                    #    args=(label, conf)
                    #).start()

                    # Si es emoc. depresiva, aumentamos contador
                    if dom in self.depressive_emotions:
                        label_persona = recognize_face(face_bgr)
                        # 1) Guardar la imagen y obtener la ruta
                        bgr_para_guardar = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)                        
                        ruta_imagen = guardar_imagen_emocion(bgr_para_guardar, label,label_persona)  # o la función que uses para guardar

                        # 2) Lanzar hilos para insertar en BD con la ruta
                        threading.Thread(
                            target=guardar_resultado_con_imagen,
                            args=(label, conf, ruta_imagen, label_persona),
                            daemon=True
                        ).start()
                        # obtenemos la etiqueta en español
                        label = self.depressive_emotions[dom]
                        # incrementamos el contador
                        self.emotions_count [label] += 1
                        print(f"[DeepFace] | Emoción depresiva detectada: {dom} ({label}) con {conf*100:.2f}%")
                    else:
                        # Si es happy/surprise/neutral, contamos también
                        label_pos = self.all_list_emotions[dom]      # p.ej. 'Feliz', 'Sorpresa' o 'Neutral'
                        self.emotions_count[label_pos] += 1

                #if dom in self.depressive_emotions:
                #    # obtenemos la etiqueta en español
                #    label = self.depressive_emotions[dom]
                #    # incrementamos el contador
                #    self.emotions_count [label] += 1
                #    print(f"[DeepFace] | Emoción depresiva detectada: {dom} ({label}) con {conf*100:.2f}%")
                # if emotion in ['sad', 'fear', 'angry', 'disgust']:
                #         threading.Thread(target=verificar_y_enviar_correo, args=(emotion, porcentaje)).start()

            # …aquí tu lógica de umbrales, dibujado de rectángulos y texto…
            # Imaginemos que pintas sobre `img` y lo guardas en `out`
            out = img.copy()

        except Exception:
            out = img  # si falla, mostramos la entrada
        finally:
            # Guardamos para que la UI lo pinte
            self.last_frame = out
            self.processing = False

    def closeEvent(self, event):
        self.cap.release()
        self.timer.stop()
        event.accept()

class FaceRegistrationDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Registrar Nuevo Rostro")
        self.cap = cv2.VideoCapture(0)
        self.face_detector = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        )

        # Widgets
        self.video_label    = QLabel()
        self.name_input     = QLineEdit()
        self.name_input.setPlaceholderText("Nombre de la persona")
        self.capture_btn    = QPushButton("Capturar (5 s.)")
        self.capture_btn.clicked.connect(self.start_timed_capture)
        self.finish_btn     = QPushButton("Finalizar")
        self.finish_btn.clicked.connect(self.stop_timed_capture)
        self.finish_btn.setEnabled(False)

        layout = QVBoxLayout()
        layout.addWidget(self.video_label)
        layout.addWidget(self.name_input)
        btns = QHBoxLayout()
        btns.addWidget(self.capture_btn)
        btns.addWidget(self.finish_btn)
        layout.addLayout(btns)
        self.setLayout(layout)

        # Timer de refresco de video
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.timer.start(30)

        # Timer para capturas periódicas
        self.capture_timer = QTimer(self)
        self.capture_timer.setInterval(500)  # cada 500 ms
        self.capture_timer.timeout.connect(self.capture_current_face)

        # Variables de estado
        self.capturing     = False
        self.capture_count = 0

    def start_timed_capture(self):
        nombre = self.name_input.text().strip()
        if not nombre:
            QMessageBox.warning(self, "Error", "Debes ingresar un nombre")
            return

        # Preparar estado
        self.capturing     = True
        self.capture_count = 0
        self.capture_btn.setEnabled(False)
        self.name_input.setEnabled(False)
        self.finish_btn.setEnabled(True)

        # Arranca capturas cada 500 ms y para tras 5 s
        self.capture_timer.start()
        QTimer.singleShot(5000, self.stop_timed_capture)

    def capture_current_face(self):
        ret, frame = self.cap.read()
        if not ret:
            return

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_detector.detectMultiScale(gray, 1.3, 5)
        # Si no hay ningun rectángulo, detectMultiScale devuelve un tuple vacío o un array vacío
        if len(faces) == 0:
            return

        x, y, w, h = faces[0]
        face_img = frame[y:y+h, x:x+w]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        nombre = self.name_input.text().strip()
        folder = "rostros_registrados"
        os.makedirs(folder, exist_ok=True)
        path = os.path.join(folder, f"{nombre}_{timestamp}.jpg")
        cv2.imwrite(path, face_img)

        # Guarda en BD
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("INSERT OR IGNORE INTO personas (nombre) VALUES (?)", (nombre,))
        c.execute("SELECT id FROM personas WHERE nombre = ?", (nombre,))
        persona_id = c.fetchone()[0]
        c.execute(
            "INSERT INTO rostros_persona (persona_id, image_path) VALUES (?, ?)",
            (persona_id, path)
        )
        conn.commit()
        conn.close()

        # Actualiza embedding en memoria
        rep = DeepFace.represent(img_path=path, model_name='Facenet',
                                 enforce_detection=False)[0]['embedding']
        vec = np.array(rep, dtype=float)
        vec = vec / np.linalg.norm(vec) if np.linalg.norm(vec)>0 else vec
        known_embeddings[nombre] = (
            known_embeddings.get(nombre, vec) + vec
        ) / 2  # simple promedio incremental

        self.capture_count += 1
        print(f"[Registro] Captura #{self.capture_count} guardada: {path}")

    def stop_timed_capture(self):
        if not self.capturing:
            return
        self.capturing = False
        self.capture_timer.stop()
        self.capture_btn.setEnabled(True)
        self.name_input.setEnabled(True)
        self.finish_btn.setEnabled(False)
        QMessageBox.information(
            self, "Listo",
            f"Se guardaron {self.capture_count} capturas para «{self.name_input.text()}»"
        )

    def update_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            return
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_detector.detectMultiScale(gray, 1.3, 5)
        for (x,y,w,h) in faces:
            cv2.rectangle(frame, (x,y), (x+w, y+h), (0,255,0), 2)
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h,w,_ = rgb.shape
        qimg = QImage(rgb.data, w, h, 3*w, QImage.Format_RGB888)
        self.video_label.setPixmap(QPixmap.fromImage(qimg))

    def closeEvent(self, event):
        if hasattr(self, 'cap') and self.cap.isOpened():
            self.cap.release()
        self.timer.stop()
        self.capture_timer.stop()
        event.accept()

def mostrar_aplicacion():
    # init_db()
    app = QApplication(sys.argv)
    window = DepressionDetector()
    window.show()
    sys.exit(app.exec_())

def load_known_faces():
    known_embeddings.clear()
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
      SELECT p.nombre, rp.image_path
      FROM personas p
      JOIN rostros_persona rp ON rp.persona_id = p.id
    ''')
    rows = c.fetchall()
    conn.close()

    # 1) Agrupamos todas las rutas por nombre
    temp = defaultdict(list)
    for nombre, path in rows:
        rep = DeepFace.represent(
            img_path=path,
            model_name='Facenet',
            enforce_detection=False,
            detector_backend='mtcnn'
        )[0]['embedding']
        temp[nombre].append(np.array(rep, dtype=float))

    # 2) Aquí promediamos y NORMALIZAMOS cada embedding
    for nombre, reps in temp.items():
        # 2.1) Promedio elemento a elemento
        avg = np.mean(reps, axis=0)

        # 2.2) L2-normalización: que la norma del vector sea 1
        norm = np.linalg.norm(avg)
        if norm > 0:
            avg = avg / norm
        print(f"[load_known_faces] {nombre}: {len(reps)} muestras → norma original={norm:.4f}")

        # 2.3) Guardamos el vector normalizado
        known_embeddings[nombre] = avg

    print(f"[load_known_faces] Total personas cargadas: {len(known_embeddings)}")

if __name__ == "__main__":
    init_db()    
    # guardar_resultado_sqlitetest()
    #mostrar_login()
    load_known_faces()
    mostrar_aplicacion()   
